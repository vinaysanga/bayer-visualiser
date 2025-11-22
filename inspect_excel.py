import pandas as pd
import openpyxl

file_path = 'data/bayer_data.xlsx'

print(f"--- Inspecting {file_path} for Prompts ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames[:3]: # Check first 3 sheets
        print(f"\nSheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # 1. Check first few cells for long text
        print("  Checking first 5 rows for potential prompt text:")
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=5):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and len(val) > 20:
                    print(f"    Cell {cell.coordinate}: {val[:100]}...")

        # 2. Check for images/drawings (Text Boxes are often shapes)
        # openpyxl support for reading text from shapes is limited, but we can see if they exist.
        # Note: 'ws.drawings' usually contains images. Comments are 'ws.comments'.
        # Text boxes might be in specific XML parts, hard to read with standard openpyxl.
        
        # However, sometimes "Text Box" means a merged cell or just a cell with a label "Prompt:"
        
except Exception as e:
    print(f"Error with openpyxl: {e}")

print("\n--- Pandas Head Check ---")
try:
    # Read with header=None to see everything
    df = pd.read_excel(file_path, sheet_name=0, header=None, nrows=5)
    print(df.to_string())
except Exception as e:
    print(f"Error with pandas: {e}")
